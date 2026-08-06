"""Historique des matchs de tennis, publie par tennis-data.co.uk.

Un classeur par saison et par circuit, en telechargement libre. Ce n'est pas une
API : c'est un fichier, traverse avec le meme retry, le meme timeout et la meme
journalisation que le reste — comme les pages HTML de Tennis Abstract.

Trois faits verifies sur la source, tous contraignants :

- **HTTPS ne repond pas**, seul `http://` sert le fichier. Aucun secret ne
  transite, aucune donnee n'est envoyee : la requete est un GET anonyme.
- Le `robots.txt` n'interdit que `/stuff/` et les saisons 2000 a 2005. Les
  saisons recentes sont explicitement autorisees — ne rien demander d'autre.
- Aucun quota, aucune cle : **rien n'est ecrit dans `api_usage`**, qui ne compte
  que des credits. Meme regle que les classements Elo.

Le classeur porte huit colonnes de **cotes de cloture** (B365, Pinnacle, Max,
Avg, Betfair). Elles sont ecartees ici, a la lecture : ce sont les prix de
fermeture du marche, donc la matiere premiere d'un calcul de CLV et de value, que
la section 9 de SPEC.md interdit. Les ecarter au parsing plutot qu'au rendu
garantit qu'elles n'entrent jamais en base et ne peuvent pas ressortir.
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

from .base import BaseHTTPClient, ProviderError

logger = logging.getLogger(__name__)

PROVIDER = "tennisdata"
BASE_URL = "http://www.tennis-data.co.uk"

#: Circuits publies, et le fragment d'URL de chacun. Le circuit feminin ajoute
#: un « w » a l'annee : `/2026w/2026.xlsx`.
TOURS = {"atp": "", "wta": "w"}

#: Sans User-Agent de navigateur, certains hebergeurs refusent le
#: telechargement. Meme precaution que pour Tennis Abstract.
USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/126.0 Safari/537.36"
)

#: Colonnes retenues. Tout le reste du classeur est ignore, et **les colonnes de
#: cotes ne figurent pas ici** : c'est la seule barriere, et elle est en amont.
COLUMNS = (
    "Date",
    "Tournament",
    "Location",
    "Series",
    "Court",
    "Surface",
    "Round",
    "Winner",
    "Loser",
    "Comment",
)

#: Colonnes de sets, par paire (vainqueur, perdant). Un match en trois sets
#: gagnants en remplit jusqu'a cinq.
SET_COLUMNS = (("W1", "L1"), ("W2", "L2"), ("W3", "L3"), ("W4", "L4"), ("W5", "L5"))

#: Interdites en base, verifiees par un test. La liste existe pour etre citee
#: dans ce test : une colonne de cotes ajoutee par la source ne doit pas se
#: glisser dans `COLUMNS` a la faveur d'une relecture distraite.
ODDS_COLUMNS = frozenset(
    {"B365W", "B365L", "PSW", "PSL", "MaxW", "MaxL", "AvgW", "AvgL", "BFEW", "BFEL"}
)


@dataclass
class RawMatch:
    """Une ligne du classeur, reduite a ce qu'un angle sportif utilise."""

    played_on: str
    tournament: str
    location: str
    series: str
    court: str
    surface: str
    round: str
    winner: str
    loser: str
    score: str
    comment: str


def _as_date(value: Any) -> str:
    """Date d'un match en ISO. Vide si la cellule n'en porte pas une.

    `openpyxl` rend un `datetime` quand la cellule est formatee en date, ce qui
    est le cas ici — le nombre de serie Excel brut (`46026`) n'apparait que si le
    format est perdu. Les deux sont geres : une date illisible fait sauter la
    ligne, jamais tomber la collecte.
    """
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, int | float):
        # Epoque Excel : le jour 1 est le 1er janvier 1900, avec le decalage
        # historique du 29 fevrier 1900 qui n'a jamais existe.
        try:
            return date.fromordinal(date(1899, 12, 30).toordinal() + int(value)).isoformat()
        except (ValueError, OverflowError):
            return ""
    return ""


def _score(cells: dict[str, Any]) -> str:
    """`6-4 3-6 7-5` a partir des colonnes de sets. Les sets absents sont omis."""
    sets = []
    for gagnant, perdant in SET_COLUMNS:
        a, b = cells.get(gagnant), cells.get(perdant)
        if a is None or b is None:
            continue
        try:
            sets.append(f"{int(a)}-{int(b)}")
        except (TypeError, ValueError):
            continue
    return " ".join(sets)


def parse_workbook(payload: bytes) -> list[RawMatch]:
    """Lit un classeur et rend ses matchs, cotes exclues.

    Le rapprochement des colonnes se fait **par libelle d'en-tete**, jamais par
    position : la source a deja ajoute des colonnes de books au fil des ans, et
    lire la dixieme colonne parce qu'elle etait `Winner` l'an dernier finirait
    par ranger un prix dans un nom de joueur.
    """
    workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = workbook.active
    if sheet is None:
        raise ProviderError(PROVIDER, "classeur", "aucune feuille exploitable")

    matches: list[RawMatch] = []
    headers: dict[int, str] = {}
    for row in sheet.iter_rows(values_only=True):
        if not headers:
            headers = {index: str(value) for index, value in enumerate(row) if value}
            continue
        cells = {
            headers[index]: value
            for index, value in enumerate(row)
            if index in headers and value is not None
        }
        played_on = _as_date(cells.get("Date"))
        winner, loser = cells.get("Winner"), cells.get("Loser")
        if not played_on or not winner or not loser:
            # Ligne de pied de tableau, ou match sans date exploitable.
            continue
        matches.append(
            RawMatch(
                played_on=played_on,
                tournament=str(cells.get("Tournament") or "").strip(),
                location=str(cells.get("Location") or "").strip(),
                series=str(cells.get("Series") or "").strip(),
                court=str(cells.get("Court") or "").strip(),
                surface=str(cells.get("Surface") or "").strip(),
                round=str(cells.get("Round") or "").strip(),
                winner=str(winner).strip(),
                loser=str(loser).strip(),
                score=_score(cells),
                comment=str(cells.get("Comment") or "").strip(),
            )
        )
    workbook.close()
    return matches


class TennisDataClient(BaseHTTPClient):
    """Telechargement des classeurs de resultats. Aucun quota, aucune cle."""

    provider_name = PROVIDER
    base_url = BASE_URL

    async def season(self, tour: str, season: int) -> list[RawMatch]:
        """Matchs d'une saison d'un circuit, cotes exclues."""
        if tour not in TOURS:
            raise ProviderError(PROVIDER, "saison", f"circuit inconnu : {tour}")
        path = f"/{season}{TOURS[tour]}/{season}.xlsx"
        result = await self._get(path, headers={"User-Agent": USER_AGENT}, as_bytes=True)
        if not isinstance(result.data, bytes | bytearray):
            raise ProviderError(PROVIDER, path, f"corps inattendu : {type(result.data).__name__}")
        matches = parse_workbook(bytes(result.data))
        logger.info(
            "%s %s — %d matchs, %d ko, %d ms (aucun credit)",
            PROVIDER,
            path,
            len(matches),
            len(result.data) // 1024,
            result.duration_ms,
        )
        return matches
