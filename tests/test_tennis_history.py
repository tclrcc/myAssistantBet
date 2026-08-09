"""Historique des matchs de tennis : source, rapprochement des noms, rendu.

Ce qui est verifie ici et qui n'allait pas de soi : les cotes de cloture du
fichier source ne doivent **jamais** entrer en base, le rapprochement des noms ne
doit jamais attribuer a un joueur l'historique d'un autre, et un match donne sur
tapis vert n'est pas un match joue.
"""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from pathlib import Path

import httpx
import pytest
import respx
from markupsafe import escape

from myassistantbet import db
from myassistantbet.config import Settings
from myassistantbet.providers.tennisdata import (
    BASE_URL,
    ODDS_COLUMNS,
    RawMatch,
    TennisDataClient,
    parse_workbook,
)
from myassistantbet.services import tennis_history

FIXTURE = Path(__file__).parent / "fixtures" / "tennisdata_atp_2026.xlsx"
NOW = datetime(2026, 8, 6, tzinfo=UTC)
COMMENCE = "2026-08-10T13:00:00Z"


@pytest.fixture
def classeur() -> bytes:
    return FIXTURE.read_bytes()


@pytest.fixture
def api_client(http_client: httpx.AsyncClient, migrated: Settings) -> TennisDataClient:
    return TennisDataClient(http_client, migrated)


def _empty_workbook() -> bytes:
    """Un classeur valide sans aucun match : en-tetes seuls.

    Un match appartient a un seul circuit et a une seule saison, et les fichiers
    de la source ne se recouvrent pas. Servir la fixture pour les six couples
    circuit/saison mettrait donc six copies du meme match en base — le bilan des
    confrontations directes vaudrait alors « 6V-0D » sur une seule rencontre.
    """
    from io import BytesIO

    from openpyxl import Workbook, load_workbook

    source = load_workbook(FIXTURE, read_only=True)
    try:
        modele = source.active
        assert modele is not None
        entetes = [cell.value for cell in next(modele.iter_rows(max_row=1))]
    finally:
        source.close()
    vide = Workbook()
    assert vide.active is not None
    vide.active.append(entetes)
    tampon = BytesIO()
    vide.save(tampon)
    return tampon.getvalue()


def _mock_seasons(classeur: bytes) -> list[respx.Route]:
    """La fixture pour l'ATP 2026, un classeur vide pour les cinq autres.

    Rend les deux routes : ce qui se compte est la somme de leurs appels, la
    premiere ne servant qu'un seul des six couples circuit/saison.
    """
    precise = respx.get(f"{BASE_URL}/2026/2026.xlsx").mock(
        return_value=httpx.Response(
            200,
            content=classeur,
            headers={"content-type": "application/vnd.openxmlformats-officedocument"},
        )
    )
    generale = respx.get(url__regex=rf"{BASE_URL}/\d{{4}}w?/\d{{4}}\.xlsx").mock(
        return_value=httpx.Response(200, content=_empty_workbook())
    )
    return [precise, generale]


def _downloads(routes: list[respx.Route]) -> int:
    """Classeurs telecharges, toutes routes confondues."""
    return sum(route.call_count for route in routes)


def _lines(
    settings: Settings, home: str, away: str, surface: str | None = "clay"
) -> dict[str, str]:
    return dict(tennis_history.lines(home, away, surface, COMMENCE, settings))


# -- Lecture du classeur -----------------------------------------------------


def test_les_cotes_de_cloture_n_entrent_jamais_en_base(classeur: bytes, migrated: Settings) -> None:
    """Interdit n°1 de SPEC.md. Le fichier porte huit colonnes de cotes de
    fermeture — B365, Pinnacle, Max, Avg, Betfair — soit la matiere premiere d'un
    calcul de CLV et de value. Elles sont ecartees **a la lecture** et non au
    rendu : ce qui n'entre pas en base ne peut pas ressortir par accident."""
    matchs = parse_workbook(classeur)
    tennis_history.store("atp", 2026, matchs, migrated, NOW)

    colonnes = {
        row["name"] for row in db.query("PRAGMA table_info(tennis_matches)", settings=migrated)
    }
    assert not colonnes & ODDS_COLUMNS, "aucune colonne de cotes dans le schema"
    champs = {champ for match in matchs for champ in vars(match)}
    assert not champs & {colonne.casefold() for colonne in ODDS_COLUMNS}
    # Et aucune valeur de cote ne doit s'etre glissee dans un champ texte.
    lignes = db.query("SELECT * FROM tennis_matches LIMIT 5", settings=migrated)
    assert lignes and all("1.9" not in str(dict(ligne)) for ligne in lignes)


def test_une_ligne_sans_date_ni_joueurs_est_ignoree(classeur: bytes) -> None:
    """Le classeur porte une ligne de pied de tableau. La compter creerait un
    match fantome, avec un vainqueur vide."""
    matchs = parse_workbook(classeur)

    assert all(match.played_on and match.winner and match.loser for match in matchs)


def test_le_score_est_reconstruit_des_colonnes_de_sets(classeur: bytes) -> None:
    """Les sets vivent dans dix colonnes ; un match en deux sets n'en remplit que
    quatre, et les colonnes vides ne doivent pas produire de tirets orphelins."""
    matchs = {(m.winner, m.loser): m for m in parse_workbook(classeur)}

    finale = matchs[("Etcheverry T.", "Zverev A.")]
    assert finale.score == "7-5 6-4"
    abandon = matchs[("Etcheverry T. M.", "Sinner J.")]
    assert abandon.score == "6-3", "un seul set joue avant l'abandon"
    forfait = matchs[("Bautista Agut R.", "Alcaraz C.")]
    assert forfait.score == "", "un forfait n'a pas de score"


# -- Rapprochement des noms --------------------------------------------------


def test_une_identite_publiee_devient_une_cle(classeur: bytes) -> None:
    """« Bautista Agut R. » a un nom de famille en deux mots : le decouper sur le
    premier espace en ferait « agut|br »."""
    assert tennis_history.published_key("Bautista Agut R.") == "bautista agut|r"
    assert tennis_history.published_key("Etcheverry T. M.") == "etcheverry|tm"
    assert tennis_history.published_key("Sinner J.") == "sinner|j"


@respx.mock
@pytest.mark.anyio
async def test_le_decoupage_prenom_nom_n_est_jamais_devine(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """« Alex de Minaur » a pour nom « de Minaur », « Juan Manuel Cerundolo » a
    pour prenoms « Juan Manuel » : rien dans la chaine ne dit lequel des deux cas
    on lit. Tous les decoupages sont essayes, et un seul resultat est accepte."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    index = tennis_history.known_keys(migrated)

    assert tennis_history.resolve("Alex de Minaur", index) == ("de minaur|a",)
    assert tennis_history.resolve("Roberto Bautista Agut", index) == ("bautista agut|r",)


@respx.mock
@pytest.mark.anyio
async def test_deux_orthographes_du_meme_joueur_sont_reunies(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Neuf paires de ce genre existent dans les donnees reelles :
    « Etcheverry T. » et « Etcheverry T. M. » sont la meme personne. Les separer
    couperait son historique en deux ; les confondre avec un homonyme serait pire.
    Le critere est le nom identique **et** des initiales en chaine de prefixes."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    index = tennis_history.known_keys(migrated)

    assert tennis_history.resolve("Tomas Martin Etcheverry", index) == (
        "etcheverry|t",
        "etcheverry|tm",
    )


@respx.mock
@pytest.mark.anyio
async def test_deux_joueurs_de_meme_nom_ne_partagent_jamais_leur_historique(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Le piege des freres Zverev. Des initiales qui divergent — `a` et `m` — ne
    forment pas une chaine de prefixes : le doute vaut silence, et il n'existe
    ici aucune resolution manuelle pour rattraper une erreur."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    index = dict(tennis_history.known_keys(migrated))
    index["zverev"] = {"zverev|a", "zverev|m"}

    assert tennis_history.resolve("Sacha Zverev", index) == (), "deux joueurs : on refuse"
    assert tennis_history.resolve("Alexander Zverev", index) == ("zverev|a",), (
        "un seul : on accepte"
    )


def test_un_joueur_inconnu_ne_produit_aucune_ligne(migrated: Settings) -> None:
    """Base vierge : ecrire « aucun match connu » ferait chercher un probleme de
    rapprochement la ou il n'y a qu'une collecte jamais lancee."""
    assert tennis_history.lines("Jannik Sinner", "Carlos Alcaraz", "hard", COMMENCE, migrated) == []


# -- Collecte ----------------------------------------------------------------


@respx.mock
@pytest.mark.anyio
async def test_la_collecte_est_idempotente(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Relancer ne duplique rien : cle naturelle (circuit, saison, date, joueurs)."""
    _mock_seasons(classeur)

    await tennis_history.refresh(api_client, migrated, now=NOW)
    avant = db.query_one("SELECT count(*) AS n FROM tennis_matches", settings=migrated)["n"]
    await tennis_history.refresh(api_client, migrated, now=NOW, force=True)
    apres = db.query_one("SELECT count(*) AS n FROM tennis_matches", settings=migrated)["n"]

    assert avant == apres > 0


@respx.mock
@pytest.mark.anyio
async def test_une_saison_terminee_n_est_jamais_retelechargee(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Elle ne changera plus. Seule la saison en cours se rafraichit, une fois
    par jour — le fichier se remplit a mesure que les tournois se terminent."""
    routes = _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    premier = _downloads(routes)

    plus_tard = NOW + timedelta(hours=tennis_history.CURRENT_SEASON_TTL_HOURS + 1)
    rapport = await tennis_history.refresh(api_client, migrated, now=plus_tard)

    assert premier == 6, "trois saisons, deux circuits"
    assert _downloads(routes) == premier + 2, "seule la saison en cours des deux circuits"
    assert rapport.seasons == ["atp 2026", "wta 2026"]


@respx.mock
@pytest.mark.anyio
async def test_un_circuit_en_echec_n_empeche_pas_l_autre(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Une source indisponible ne coute que des lignes, jamais la collecte."""
    # Enregistree **avant** les autres : respx retient la premiere route qui
    # correspond, et l'inverse ne declencherait jamais le 404.
    respx.get(url__regex=rf"{BASE_URL}/\d{{4}}w/\d{{4}}\.xlsx").mock(
        return_value=httpx.Response(404, text="absent")
    )
    _mock_seasons(classeur)

    rapport = await tennis_history.refresh(api_client, migrated, now=NOW)

    assert not rapport.ok
    assert rapport.seasons == ["atp 2026", "atp 2025", "atp 2024"]
    assert len(rapport.errors) == 3


@respx.mock
@pytest.mark.anyio
async def test_aucun_credit_n_est_comptabilise(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Gratuit et sans cle : `api_usage` ne compte que des credits, et y ecrire un
    telechargement libre fausserait le bandeau."""
    _mock_seasons(classeur)

    await tennis_history.refresh(api_client, migrated, now=NOW)

    assert db.query("SELECT * FROM api_usage", settings=migrated) == []


# -- Rendu -------------------------------------------------------------------


@respx.mock
@pytest.mark.anyio
async def test_le_bloc_porte_les_confrontations_directes_avec_leurs_scores(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Un 3V-1D dont les trois victoires tiennent en trois sets serres ne decrit
    pas le meme rapport de forces qu'un 3V-1D en deux sets secs. Le bilan s'ecrit
    `V-D` comme partout ailleurs dans le projet, jamais `3-1` — qui se lirait
    comme un score."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)

    lignes = _lines(migrated, "Tomas Martin Etcheverry", "Alexander Zverev")
    h2h = next(value for label, value in lignes.items() if label.startswith("H2H"))
    assert "Tomas Martin Etcheverry 1V-0D" in h2h
    assert "07/26 terre 7-5 6-4" in h2h, "la date porte l'annee : trois saisons se melangeraient"


@respx.mock
@pytest.mark.anyio
async def test_un_abandon_est_dit_dans_le_score(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """« 6-3 » seul sur une confrontation directe passerait pour une donnee
    tronquee. C'est un match interrompu, et ca se lit."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)

    lignes = _lines(migrated, "Tomas Martin Etcheverry", "Jannik Sinner")
    h2h = next(value for label, value in lignes.items() if label.startswith("H2H"))
    assert "6-3 ab." in h2h


@respx.mock
@pytest.mark.anyio
async def test_un_forfait_n_est_pas_un_match_joue(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Personne n'est entre sur le court : le compter en confrontation directe
    donnerait un rapport de forces sur un match qui n'a pas eu lieu. Il reste une
    information sur la disponibilite, portee par la ligne « Abandons »."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)

    lignes = _lines(migrated, "Roberto Bautista Agut", "Carlos Alcaraz")
    # Le forfait etait leur seul match : il ne compte pas comme une rencontre, et
    # la ligne le dit sans pretendre qu'ils n'ont jamais ete tires ensemble.
    assert lignes["H2H"] == "aucun match joue depuis 2024"
    assert "forfait" in lignes["Abandons"]


@respx.mock
@pytest.mark.anyio
async def test_le_bilan_de_surface_porte_sa_fenetre(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """« 8V-3D » sur un an et sur trois ans ne disent pas la meme chose — meme
    regle que le compte a cote d'une moyenne."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)

    surface = _lines(migrated, "Jannik Sinner", "Carlos Alcaraz")["Surface"]
    assert "terre" in surface
    assert "/12m" in surface


@respx.mock
@pytest.mark.anyio
async def test_aucune_ligne_de_surface_sans_surface_renseignee(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """La surface est saisie a la main sur la competition. La deduire du libelle
    d'un tournoi serait une invention — meme regle que pour l'Elo."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)

    lignes = _lines(migrated, "Jannik Sinner", "Carlos Alcaraz", surface=None)
    assert "Surface" not in lignes
    assert "Forme" in lignes, "le reste du bloc n'en depend pas"


@respx.mock
@pytest.mark.anyio
async def test_la_forme_se_lit_dans_le_meme_sens_qu_au_football(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """La derniere lettre est le dernier match, comme « Forme 5 ». Inverser le
    sens sur un seul des deux sports serait un piege a coup sur."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)

    forme = _lines(migrated, "Jannik Sinner", "Carlos Alcaraz")["Forme"]
    lettres = forme.split("Jannik Sinner ")[1].split("/")[0]
    assert set(lettres) <= {"V", "D"}
    dernier = db.query_one(
        "SELECT winner_key FROM tennis_matches WHERE winner_key LIKE 'sinner|%' "
        "OR loser_key LIKE 'sinner|%' ORDER BY played_on DESC LIMIT 1",
        settings=migrated,
    )
    attendu = "V" if str(dernier["winner_key"]).startswith("sinner|") else "D"
    assert lettres[-1] == attendu


@respx.mock
@pytest.mark.anyio
async def test_une_saison_sans_aucun_match_n_est_pas_redemandee_sans_fin(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Piege trouve en ecrivant le test precedent : la peremption etait deduite du
    `MAX(fetched_at)` des matchs, donc une saison vide n'avait pas de date, donc
    elle passait pour jamais telechargee. En janvier, le fichier de la saison qui
    commence est justement vide : la collecte se serait relancee a chaque
    enrichissement, sans fin et sans que rien ne le dise."""
    routes = _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    premier = _downloads(routes)

    await tennis_history.refresh(api_client, migrated, now=NOW)

    assert _downloads(routes) == premier, "rien n'est frais a redemander"
    vides = db.query(
        "SELECT tour, season FROM tennis_history_state WHERE matches = 0", settings=migrated
    )
    assert len(vides) == 5, "cinq couples circuit/saison sans match, tous memorises"


def test_le_prompt_encadre_les_bilans_de_tennis(migrated: Settings) -> None:
    """Le garde-fou compte autant que la donnee. Un bilan de confrontations est une
    frequence passee, comme les fractions du football : jamais rapprochee d'une
    cote. Et une ligne absente n'est pas un joueur sans passe.

    Le lot porte de vrais matchs : le preambule ne documente que les lignes
    presentes, et un lot sans historique n'a pas a payer leur mode d'emploi."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)

    body = build_prompt(session, settings=migrated, now=NOW).body

    assert "ne les rapproche jamais d'une cote" in body
    assert "tapis vert n'entre dans aucun de ces comptes" in body
    assert "n'est pas un joueur sans passé" in body
    assert "même sens de lecture que « Forme 5 »" in body


# -- Ce qui s'est passe dans ce tournoi --------------------------------------


def _competition(settings: Settings, cle: str, tournois: str | None) -> int:
    """Renseigne la correspondance d'un tournoi et rend son identifiant."""
    row = db.query_one(
        "SELECT id FROM competitions WHERE oddsapi_key = ?", (cle,), settings=settings
    )
    assert row is not None
    db.execute(
        "UPDATE competitions SET tennisdata_tournaments = ? WHERE id = ?",
        (tournois, row["id"]),
        settings=settings,
    )
    return int(row["id"])


@respx.mock
@pytest.mark.anyio
async def test_le_palmares_distingue_une_finale_gagnee_d_une_finale_perdue(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """C'est l'erreur la plus visible que la ligne pourrait commettre : le rang du
    tour ne dit pas l'issue. Une finale gagnee vaut « vainqueur », perdue
    « finaliste »."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    competition = _competition(migrated, "tennis_atp_french_open", "Generali Open")

    lignes = dict(
        tennis_history.lines(
            "Tomas Martin Etcheverry", "Alexander Zverev", None, COMMENCE, migrated, competition
        )
    )

    assert "Tomas Martin Etcheverry vainqueur 2026" in lignes["Palmares"]
    assert "Alexander Zverev finaliste 2026" in lignes["Palmares"]


@respx.mock
@pytest.mark.anyio
async def test_les_confrontations_dans_ce_tournoi_portent_leur_tour(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Un huitieme de finale et une finale dans le meme tournoi ne se valent pas,
    et c'est precisement ce que la question « ici » cherche."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    competition = _competition(migrated, "tennis_atp_french_open", "Generali Open")

    lignes = dict(
        tennis_history.lines(
            "Tomas Martin Etcheverry", "Alexander Zverev", None, COMMENCE, migrated, competition
        )
    )

    assert lignes["H2H ici"] == (
        "Tomas Martin Etcheverry 1V-0D · 07/26 finale (Tomas Martin Etcheverry)"
    )


@respx.mock
@pytest.mark.anyio
async def test_sans_correspondance_de_tournoi_aucune_ligne_ici(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """La source nomme les tournois par leur sponsor : « ABN AMRO World Tennis
    Tournament » pour Rotterdam. Rien ne se deduit d'un libelle, et un palmares
    emprunte a un autre tournoi serait pire qu'une ligne absente."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    competition = _competition(migrated, "tennis_atp_french_open", None)

    lignes = dict(
        tennis_history.lines(
            "Tomas Martin Etcheverry", "Alexander Zverev", None, COMMENCE, migrated, competition
        )
    )

    assert "Palmares" not in lignes
    assert "H2H ici" not in lignes
    assert "Forme" in lignes, "le reste du bloc n'en depend pas"


@respx.mock
@pytest.mark.anyio
async def test_un_tournoi_renomme_par_son_sponsor_reste_le_meme_tournoi(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """La source porte deja deux orthographes pour l'epreuve de Houston. Plusieurs
    noms se separent par `|` : les separer couperait un palmares en deux."""
    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    seul = _competition(migrated, "tennis_atp_french_open", "Generali Open")
    un_nom = dict(
        tennis_history.lines(
            "Tomas Martin Etcheverry", "Alexander Zverev", None, COMMENCE, migrated, seul
        )
    )
    _competition(migrated, "tennis_atp_french_open", "Generali Open|Nordea Open")
    deux_noms = dict(
        tennis_history.lines(
            "Tomas Martin Etcheverry", "Alexander Zverev", None, COMMENCE, migrated, seul
        )
    )

    # Le Generali Open porte sa finale, le Nordea Open sa victoire sur abandon.
    assert "Tomas Martin Etcheverry vainqueur 2026, 1V-0D" in un_nom["Palmares"]
    assert "Tomas Martin Etcheverry vainqueur 2026, 2V-0D" in deux_noms["Palmares"]


def test_la_table_de_correspondance_couvre_nos_tournois(migrated: Settings) -> None:
    """Le seed de la migration 020 et la table appliquee a la synchronisation
    doivent dire la meme chose : deux sources de verite finiraient par diverger."""
    from myassistantbet.services.competitions import TENNISDATA_TOURNAMENTS

    rows = db.query(
        "SELECT c.oddsapi_key, c.tennisdata_tournaments AS noms FROM competitions c "
        "JOIN sports s ON s.id = c.sport_id WHERE s.key = 'tennis'",
        settings=migrated,
    )
    assert rows, "le seed cree bien des competitions de tennis"
    for row in rows:
        attendu = TENNISDATA_TOURNAMENTS.get(row["oddsapi_key"])
        assert row["noms"] == attendu, row["oddsapi_key"]


def test_le_prompt_dit_ce_que_l_absence_des_lignes_ici_signifie(migrated: Settings) -> None:
    """Leur absence ne dit rien du passe des joueurs sur place : elle dit que le
    rattachement du tournoi manque. Confondre les deux ferait conclure d'un
    silence.

    Cette note explique une **absence** : elle ne peut donc pas se garder sur la
    presence de « Palmares ». Elle suit le sort du bloc d'historique, qui n'a de
    raison d'exister que si le lot porte des lignes d'historique."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)

    body = build_prompt(session, settings=migrated, now=NOW).body

    assert "« finaliste » veut dire finale" in body
    assert "elle dit que le rattachement manque" in body


# -- Un seul assembleur de bloc ----------------------------------------------


@respx.mock
@pytest.mark.anyio
async def test_la_fiche_d_un_match_de_tennis_porte_le_meme_bloc_que_le_prompt(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """La fiche d'un match de tennis affichait un bloc CONTEXTE **vide** : l'Elo,
    le repos et l'historique n'existaient que dans le prompt. Deux assemblages
    paralleles ont diverge deux fois — c'est desormais le meme appel des deux
    cotes, et ce test compare les deux resultats."""
    from fastapi.testclient import TestClient

    from myassistantbet.main import app
    from myassistantbet.services import session as session_service

    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    competition = _competition(migrated, "tennis_atp_french_open", "Generali Open")
    db.execute(
        "INSERT INTO events (id, sport_id, competition_id, home, away, commence_time, source, "
        "created_at) SELECT 1, sport_id, ?, 'Tomas Martin Etcheverry', 'Alexander Zverev', ?, "
        "'api', ? FROM competitions WHERE id = ?",
        (competition, COMMENCE, db.utcnow(), competition),
        settings=migrated,
    )

    attendu = session_service.context_block(
        1,
        "Tomas Martin Etcheverry",
        "Alexander Zverev",
        COMMENCE,
        "tennis",
        oddsapi_key="tennis_atp_french_open",
        surface=None,
        competition_id=competition,
        settings=migrated,
    )
    assert any(label == "Palmares" for label, _ in attendu), "le bloc porte bien des lignes"

    with TestClient(app) as client:
        page = client.get("/events/1")

    for label, value in attendu:
        assert label in page.text, f"la fiche doit porter la ligne « {label} »"
        # La fiche passe par Jinja, qui echappe : une valeur portant une
        # apostrophe echouerait ici pour une raison qui n'a rien a voir avec la
        # divergence que ce test surveille, et se ferait « corriger » en
        # affaiblissant l'assertion.
        assert str(escape(value.split(" · ")[0])) in page.text


@respx.mock
@pytest.mark.anyio
async def test_la_fiche_montre_le_detail_que_la_ligne_forme_ne_peut_pas_dire(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """`VVDVDDVVVD` se lit comme un joueur irregulier ; le detail montre d'ou
    viennent les defaites. Sur l'ecran et non dans le prompt : dix rencontres par
    joueur avec adversaire, score, tournoi et tour couteraient cinq cents
    caracteres par bloc."""
    from fastapi.testclient import TestClient

    from myassistantbet.main import app

    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    competition = _competition(migrated, "tennis_atp_french_open", "Generali Open")
    db.execute(
        "INSERT INTO events (id, sport_id, competition_id, home, away, commence_time, source, "
        "created_at) SELECT 1, sport_id, ?, 'Tomas Martin Etcheverry', 'Alexander Zverev', ?, "
        "'api', ? FROM competitions WHERE id = ?",
        (competition, COMMENCE, db.utcnow(), competition),
        settings=migrated,
    )

    with TestClient(app) as client:
        page = client.get("/events/1").text

    assert "Derniers matchs joués" in page
    # Le detail : adversaire, score, tournoi et tour, pas seulement une lettre.
    assert "Zverev A." in page
    assert "7-5 6-4" in page
    assert "The Final" in page
    # Et le prompt, lui, n'en porte rien : son budget est en tokens.
    from myassistantbet.services.prompt import build_prompt

    db.execute(
        "INSERT INTO sessions (id, label, created_at) VALUES (9, 'test', ?)",
        (db.utcnow(),),
        settings=migrated,
    )
    db.execute("INSERT INTO session_events (session_id, event_id) VALUES (9, 1)", settings=migrated)
    corps = build_prompt(9, settings=migrated).body
    assert "7-5 6-4" in corps, "le score figure dans la ligne H2H ici"
    assert "Derniers matchs" not in corps, "mais pas la liste detaillee"


@respx.mock
@pytest.mark.anyio
async def test_un_joueur_non_rapproche_le_dit_sur_la_fiche(
    api_client: TennisDataClient, migrated: Settings, classeur: bytes
) -> None:
    """Sans ce message, l'absence de tableau passerait pour une panne."""
    from fastapi.testclient import TestClient

    from myassistantbet.main import app

    _mock_seasons(classeur)
    await tennis_history.refresh(api_client, migrated, now=NOW)
    competition = _competition(migrated, "tennis_atp_french_open", "Generali Open")
    db.execute(
        "INSERT INTO events (id, sport_id, competition_id, home, away, commence_time, source, "
        "created_at) SELECT 1, sport_id, ?, 'Tomas Martin Etcheverry', 'Joueur Inconnu', ?, "
        "'api', ? FROM competitions WHERE id = ?",
        (competition, COMMENCE, db.utcnow(), competition),
        settings=migrated,
    )

    with TestClient(app) as client:
        page = client.get("/events/1").text

    assert "Aucun match dans l'historique collecté pour ce joueur." in page


# -- Dates hors saison -------------------------------------------------------


def _raw(played_on: str, tournament: str = "Iasi Open") -> RawMatch:
    return RawMatch(
        played_on=played_on,
        tournament=tournament,
        location="Iasi",
        series="WTA250",
        court="Outdoor",
        surface="Clay",
        round="The Final",
        winner="Sherif M.",
        loser="Badosa P.",
        score="6-4 4-0",
        comment="Retired",
    )


def test_la_saison_ouvre_en_decembre_de_l_annee_precedente() -> None:
    """Le garde-fou evident — « l'annee de la date egale la saison » — serait
    faux. Releve en base : le fichier 2025 porte 69 matchs joues du 29 au 31
    decembre 2024, celui de 2024 onze matchs du 31 decembre 2023. Les jeter
    amputerait chaque saison de son ouverture."""
    assert tennis_history.in_season("2024-12-29", 2025) is True
    assert tennis_history.in_season("2024-12-31", 2025) is True
    assert tennis_history.in_season("2025-01-01", 2025) is True
    assert tennis_history.in_season("2025-11-16", 2025) is True
    # Novembre de l'annee precedente, en revanche, est la saison d'avant.
    assert tennis_history.in_season("2024-11-16", 2025) is False


def test_une_date_posterieure_a_la_saison_est_une_coquille() -> None:
    """Le fichier 2026 datait une finale de l'Iasi Open du 20 juillet 2029."""
    assert tennis_history.in_season("2029-07-20", 2026) is False
    assert tennis_history.in_season("pas une date", 2026) is False


def test_une_date_hors_saison_n_entre_pas_en_base(migrated: Settings) -> None:
    """Le degat qu'elle fait est invisible : posterieure a tout match analyse,
    elle sort de chaque fenetre de forme, de surface et de H2H. Le match ne
    s'affiche donc nulle part, et rien ne signale le trou."""
    ecrits = tennis_history.store(
        "wta", 2026, [_raw("2026-07-20"), _raw("2029-07-20", "Iasi Open bis")], migrated, NOW
    )

    assert ecrits == 1
    dates = [
        row["played_on"]
        for row in db.query("SELECT played_on FROM tennis_matches", settings=migrated)
    ]
    assert dates == ["2026-07-20"]


def test_le_compte_de_la_collecte_ne_retient_que_les_lignes_gardees(migrated: Settings) -> None:
    """`tennis_history_state.matches` sert a savoir ce qui a ete collecte : y
    compter une ligne jetee ferait chercher en base un match qui n'y est pas."""
    tennis_history.store("wta", 2026, [_raw("2026-07-20"), _raw("2029-07-20")], migrated, NOW)

    etat = db.query_one(
        "SELECT matches FROM tennis_history_state WHERE tour = 'wta' AND season = 2026",
        settings=migrated,
    )
    assert etat["matches"] == 1


def test_la_migration_purge_les_dates_deja_ecrites(migrated: Settings) -> None:
    """Le garde-fou de `store()` ne nettoie pas le passe : une saison terminee
    n'est jamais retelechargee, et la saison en cours ne reecrit pas une ligne
    qu'elle ne renvoie plus. La migration 021 s'en charge une fois.

    Elle rejoue **le texte du fichier**, pas une copie de la regle : la version
    SQL et `in_season()` sont deux ecritures du meme critere, et rien d'autre ne
    les empeche de diverger.
    """
    from myassistantbet.config import PACKAGE_DIR

    lignes = [
        ("wta", 2026, "2026-07-20", "Iasi Open"),  # dans sa saison
        ("wta", 2026, "2029-07-20", "Iasi Open bis"),  # coquille de la source
        ("atp", 2025, "2024-12-29", "United Cup"),  # ouverture de saison, valide
        ("atp", 2025, "2024-11-29", "Tournoi d'avant"),  # saison precedente
    ]
    for tour, season, played_on, tournament in lignes:
        db.execute(
            "INSERT INTO tennis_matches (tour, season, played_on, tournament, winner, loser, "
            "winner_key, loser_key, fetched_at) VALUES (?, ?, ?, ?, 'A', 'B', ?, 'b|', ?)",
            (tour, season, played_on, tournament, f"a|{played_on}", db.utcnow()),
            settings=migrated,
        )

    sql = (PACKAGE_DIR / "migrations" / "021_dates_hors_saison.sql").read_text(encoding="utf-8")
    with db.connect(migrated) as conn:
        conn.execute(sql)

    restant = [
        row["played_on"]
        for row in db.query(
            "SELECT played_on FROM tennis_matches ORDER BY played_on", settings=migrated
        )
    ]
    assert restant == ["2024-12-29", "2026-07-20"]


def _lot(
    settings: Settings,
    sport: str,
    home: str = "A",
    away: str = "B",
    commence: str = "2099-01-01T18:00:00Z",
) -> int:
    """Une session portant un match de ce sport, et son identifiant.

    Les noms et l'horaire se choisissent parce que le preambule ne documente pas
    seulement les **sports** du lot : il ne documente que les **lignes de
    contexte reellement presentes**. Un mode d'emploi ne se teste donc que sur un
    lot qui porte la ligne, et l'affiche « A – B » ne porte rien.

    Le preambule du prompt ne documente que les sports **presents dans le lot** :
    une session de football n'a pas a payer les quarante lignes d'explication du
    tennis. Ces tests portent donc sur un lot du bon sport — sur une session
    vide, aucun garde-fou ne se rendrait, et pour cause.
    """
    row = db.query_one(f"SELECT id FROM sports WHERE key = '{sport}'", settings=settings)
    db.execute(
        "INSERT INTO events (id, sport_id, home, away, commence_time, source, created_at) "
        "VALUES (900, ?, ?, ?, ?, 'api', ?)",
        (row["id"], home, away, commence, db.utcnow()),
        settings=settings,
    )
    db.execute(
        "INSERT INTO sessions (id, label, created_at) VALUES (1, 'test', ?)",
        (db.utcnow(),),
        settings=settings,
    )
    db.execute(
        "INSERT INTO session_events (session_id, event_id) VALUES (1, 900)", settings=settings
    )
    return 1


# -- Forme d'un match : profil de jeux, marge, niveau des adversaires ---------


def _joue(
    settings: Settings,
    played_on: str,
    score: str,
    *,
    won: bool = True,
    joueur: str = "lehecka|j",
    contre: str = "kopriva|v",
    tour: str = "atp",
    series: str = "ATP500",
    comment: str = "Completed",
) -> None:
    """Un match en base, du point de vue du joueur interroge.

    Le score est toujours ecrit gagnant d'abord, comme le fichier source le
    publie : c'est ce sens qui rend la marge negative pour le perdant.
    """
    winner, loser = (joueur, contre) if won else (contre, joueur)
    db.execute(
        "INSERT INTO tennis_matches (tour, season, played_on, tournament, series, surface, "
        "round, winner, loser, winner_key, loser_key, score, comment, fetched_at) "
        "VALUES (?, ?, ?, 'Tournoi', ?, 'Hard', '1st Round', ?, ?, ?, ?, ?, ?, ?)",
        (
            tour,
            int(played_on[:4]),
            played_on,
            series,
            winner.split("|")[0].title(),
            loser.split("|")[0].title(),
            winner,
            loser,
            score,
            comment,
            db.utcnow(),
        ),
        settings=settings,
    )


def _serie(settings: Settings, scores: list[tuple[str, bool]], **kwargs: object) -> None:
    """Une suite de matchs, un par jour, du plus ancien au plus recent."""
    for index, (score, won) in enumerate(scores):
        _joue(settings, f"2026-07-{index + 1:02d}", score, won=won, **kwargs)  # type: ignore[arg-type]


def test_le_profil_donne_la_forme_d_un_match(migrated: Settings) -> None:
    """« Usure » donne une moyenne, qui dit le temps passe sur le court. Elle ne
    dit pas si ce joueur produit des matchs serres ou des matchs a sens unique,
    et c'est la question des marches de jeux — que rien n'eclairait."""
    _serie(
        migrated,
        [
            ("7-6 6-7 7-6", True),
            ("7-6 7-6", True),
            ("6-4 6-4", True),
            ("6-1 6-2", True),
            ("6-3 6-4", False),
        ],
    )

    lignes = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")
    profil = lignes["Profil"]

    # 39, 26, 20, 15 et 19 jeux : la mediane vaut 20 la ou la moyenne vaut 23.8.
    assert "Jiri Lehecka med 20 jeux (15-39)" in profil
    assert "TB 2/5" in profil, "deux matchs contenaient un tie-break, pas trois sets"
    assert "2 sets 4/5" in profil


def test_le_grand_chelem_masculin_sort_du_profil_mais_reste_dans_l_usure(
    migrated: Settings,
) -> None:
    """Quarante jeux sont ordinaires au meilleur des cinq sets et exceptionnels
    ailleurs : les melanger fait lire un joueur de trois sets comme un marathonien.
    « Usure » les garde, elle — cinq sets fatiguent vraiment."""
    _serie(migrated, [("6-4 6-4", True)] * 5)
    _joue(migrated, "2026-07-20", "7-6 6-7 7-6 6-7 7-5", won=True, series="Grand Slam")

    lignes = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")

    assert "med 20 jeux (20-20)" in lignes["Profil"], "le Bo5 ne doit pas peser"
    assert "sur 5" in lignes["Profil"] or "/5" in lignes["Profil"]
    assert "sur 6" in lignes["Usure"], "l'usure, elle, compte les cinq sets"


def test_le_meme_grand_chelem_compte_au_feminin(migrated: Settings) -> None:
    """La colonne `series` est vide cote WTA, et tout s'y joue en trois sets :
    ecarter les Grands Chelems des deux cotes viderait la ligne d'un tableau
    entier."""
    _serie(migrated, [("6-4 6-4", True)] * 5, joueur="swiatek|i", tour="wta", series="")
    _joue(
        migrated,
        "2026-07-20",
        "6-4 6-4",
        joueur="swiatek|i",
        tour="wta",
        series="Grand Slam",
    )

    lignes = _lines(migrated, "Iga Swiatek", "Vit Kopriva")

    assert "2 sets 6/6" in lignes["Profil"]


def test_la_marge_separe_les_victoires_des_defaites(migrated: Settings) -> None:
    """Melangees en une moyenne unique, elles s'annulent : ce joueur gagne de
    huit jeux et perd de huit, et une moyenne a zero le decrirait comme un joueur
    qui joue serre. C'est la grandeur du handicap jeux."""
    _serie(
        migrated,
        [
            ("6-1 6-2", True),
            ("6-2 6-1", True),
            ("6-2 6-1", False),
            ("6-1 6-2", False),
            ("6-3 6-3", True),
        ],
    )

    marge = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")["Marge"]

    assert "+8.0 en V/3" in marge
    assert "-9.0 en D/2" in marge


def test_sous_cinq_matchs_aucune_forme_de_match(migrated: Settings) -> None:
    """Une mediane sur quatre matchs decrit une semaine, pas une tendance, et la
    lire comme telle est pire que de ne rien lire. Meme regle que le profil
    corners du football."""
    _serie(migrated, [("6-4 6-4", True)] * 4)

    lignes = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")

    assert "Profil" not in lignes
    assert "Marge" not in lignes
    assert "Forme" in lignes, "la forme, elle, se rend des le premier match"


def test_le_niveau_des_adversaires_corrige_la_lecture_de_la_forme(migrated: Settings) -> None:
    """`DDDDDVVVVV` traite une victoire sur le 150e comme une victoire sur le 5e.
    L'Elo des adversaires est deja en base pour la ligne « Elo » : la ligne ne
    coute aucun appel."""
    _serie(migrated, [("6-4 6-4", True)] * 3 + [("6-4 6-4", False)] * 2)
    _joue(migrated, "2026-07-10", "6-4 6-4", won=True, contre="musetti|l")
    for player, elo in (("Vit Kopriva", 1700.0), ("Lorenzo Musetti", 1950.0)):
        db.execute(
            "INSERT INTO tennis_elo (tour, normalized, player, elo, fetched_at) "
            "VALUES ('atp', ?, ?, ?, ?)",
            (player.casefold(), player, elo, db.utcnow()),
            settings=migrated,
        )

    niveau = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")["Niveau adv."]

    assert "Jiri Lehecka adv. Elo moy 1742/6" in niveau
    assert "meilleur battu 1950" in niveau


def test_le_meilleur_battu_ne_sort_jamais_d_une_defaite(migrated: Settings) -> None:
    """Il nomme un fait — cet adversaire-la a ete battu. Le prendre sur toutes
    les rencontres creditrait un joueur du niveau de celui qui l'a sorti."""
    _serie(migrated, [("6-4 6-4", True)] * 5)
    _joue(migrated, "2026-07-10", "6-4 6-4", won=False, contre="musetti|l")
    for player, elo in (("Vit Kopriva", 1700.0), ("Lorenzo Musetti", 1950.0)):
        db.execute(
            "INSERT INTO tennis_elo (tour, normalized, player, elo, fetched_at) "
            "VALUES ('atp', ?, ?, ?, ?)",
            (player.casefold(), player, elo, db.utcnow()),
            settings=migrated,
        )

    niveau = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")["Niveau adv."]

    assert "meilleur battu 1700" in niveau


def test_une_cle_disputee_par_deux_joueurs_du_classement_est_ecartee(migrated: Settings) -> None:
    """Garder le dernier arrive rendrait l'erreur silencieuse le jour ou deux
    homonymes apparaissent au classement. Meme regle que le rapprochement des
    noms : en cas de doute, aucune ligne."""
    _serie(migrated, [("6-4 6-4", True)] * 5)
    for player, elo in (("Vit Kopriva", 1700.0), ("Vaclav Kopriva", 1950.0)):
        db.execute(
            "INSERT INTO tennis_elo (tour, normalized, player, elo, fetched_at) "
            "VALUES ('atp', ?, ?, ?, ?)",
            (player.casefold(), player, elo, db.utcnow()),
            settings=migrated,
        )

    index = tennis_history.known_keys(migrated)
    ratings = tennis_history.ratings_by_key(index, migrated)

    assert ("atp", "kopriva|v") not in ratings
    assert "Niveau adv." not in _lines(migrated, "Jiri Lehecka", "Vit Kopriva")


def test_le_preambule_documente_les_trois_lignes(migrated: Settings) -> None:
    """Un marche ajoute sans son mode d'emploi sort en cle brute ; une ligne de
    contexte ajoutee sans le sien se lit de travers, ce qui est pire."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)
    corps = build_prompt(session, settings=migrated, now=NOW).body

    assert "« Niveau adv. »" in corps
    assert "« Profil »" in corps
    assert "« Marge »" in corps
    assert "handicap jeux" in corps
    assert "meilleur des cinq sets" in corps


# -- Jusqu'ou va l'historique ------------------------------------------------


def test_un_historique_en_retard_le_dit(migrated: Settings) -> None:
    """Le fichier source est hebdomadaire et publie apres coup : le 8 aout il
    s'arretait au 3, et aucun match du Canadian Open — commence le 4 — n'existait
    en base. Toutes les lignes tirees de l'historique s'arretaient donc avant le
    tournoi en cours sans que rien ne le dise, et le trou se lisait comme un
    rapprochement rate."""
    _serie(migrated, [("6-4 6-4", True)] * 5)

    lignes = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")

    assert lignes["Historique"] == "dernier match connu le 05/07, soit 36j avant celui-ci"


def test_un_historique_a_jour_ne_dit_rien(migrated: Settings) -> None:
    """Un fichier frais accuse deja trois a quatre jours de retard sur la
    semaine en cours : sous le seuil, la ligne se rendrait sans qu'il manque
    quoi que ce soit, et ferait douter de donnees completes."""
    _serie(migrated, [("6-4 6-4", True)] * 5)
    _joue(migrated, "2026-08-09", "6-4 6-4", won=True)

    assert "Historique" not in _lines(migrated, "Jiri Lehecka", "Vit Kopriva")


def test_le_retard_se_compte_par_circuit(migrated: Settings) -> None:
    """L'ATP et la WTA sont deux fichiers : l'un peut etre a jour quand l'autre
    ne l'est pas, et lire le plus recent des deux tairait le retard du bon."""
    _serie(migrated, [("6-4 6-4", True)] * 5, joueur="swiatek|i", tour="wta", series="")
    # Un match ATP tout frais, entre deux joueurs etrangers a cette rencontre :
    # le prendre en compte tairait le retard du circuit qui nous interesse.
    _joue(migrated, "2026-08-09", "6-4 6-4", joueur="fritz|t", contre="gea|a", tour="atp")

    lignes = _lines(migrated, "Iga Swiatek", "Vit Kopriva")

    assert lignes["Historique"] == "dernier match connu le 05/07, soit 36j avant celui-ci"


def test_la_ligne_n_affirme_jamais_que_ce_tournoi_manque(migrated: Settings) -> None:
    """Ce serait faux d'un tournoi commence avant la date de collecte, et une
    affirmation fausse dans la ligne qui sert justement a douter est le pire
    endroit ou en mettre une. Elle enonce un fait, le template en tire la suite."""
    _serie(migrated, [("6-4 6-4", True)] * 5)

    valeur = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")["Historique"]

    assert "tournoi" not in valeur
    assert valeur.startswith("dernier match connu le ")


def test_le_preambule_dit_ce_que_le_retard_implique(migrated: Settings) -> None:
    """La ligne donne une date ; c'est le preambule qui dit que « Forme » peut
    ignorer deux victoires de la semaine, et que « Parcours » comble le trou.

    Le lot porte de vrais matchs : sans la ligne, son mode d'emploi ne se rend
    pas — et c'est voulu."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)
    corps = build_prompt(session, settings=migrated, now=NOW).body

    assert "Historique" in corps, "le lot porte bien la ligne"

    assert "« Historique »" in corps
    assert "hebdomadaire" in corps
    assert "« Parcours »" in corps


def test_le_prompt_rappelle_que_le_vainqueur_n_est_pas_le_debouche_par_defaut(
    migrated: Settings,
) -> None:
    """Vingt-cinq des trente premieres selections tennis portaient sur un
    Vainqueur. La section B demandait « le marche qui traduit le mieux l'angle »
    sans jamais dire que ce marche-la ne retient qu'un nom de camp.

    Le rappel reste **sportif** : suggerer d'aller chercher un marche mieux paye
    serait raisonner sur le prix, ce qu'interdit la section 9. Un test le verifie."""
    from myassistantbet.services.prompt import build_prompt

    # Le texte est justifie a une largeur fixe : chercher une phrase entiere
    # dans le corps brut la couperait a la premiere fin de ligne.
    corps = " ".join(build_prompt(_lot(migrated, "tennis"), settings=migrated).body.split())

    assert "le plus grossier des débouchés d'une analyse" in corps
    assert "ce serait raisonner sur le prix" in corps
    assert "« Profil » et « Marge » sont là pour ça" in corps


def test_l_usure_dit_combien_de_matchs_se_jouaient_en_cinq_sets(migrated: Settings) -> None:
    """Ils restent comptes — trente-neuf jeux fatiguent autant quel que soit le
    format — mais leur nombre est dit, parce que c'est la **comparaison** qu'ils
    faussaient. Lehecka affichait 32.3 jeux/match contre 30.5 a Jodar sans que
    rien ne dise que quatre de ses dix matchs etaient un Grand Chelem."""
    _serie(migrated, [("6-4 6-4", True)] * 4)
    _joue(migrated, "2026-07-20", "7-6 6-7 7-6 6-7 7-5", won=True, series="Grand Slam")

    lignes = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")

    assert lignes["Usure"].startswith("Jiri Lehecka 28.8 jeux/match sur 5 (1 en 5 sets)")


def test_l_usure_ne_compte_pas_les_cinq_sets_quand_il_n_y_en_a_pas(migrated: Settings) -> None:
    """Une donnee absente ne produit jamais un « 0 » : c'est le cas ordinaire, et
    l'ecrire couterait une mention a chaque bloc pour ne rien apprendre."""
    _serie(migrated, [("6-4 6-4", True)] * 5)

    usure = _lines(migrated, "Jiri Lehecka", "Vit Kopriva")["Usure"]

    assert usure.startswith("Jiri Lehecka 20.0 jeux/match sur 5")
    assert "5 sets" not in usure


def test_le_preambule_dit_ce_que_le_compte_de_cinq_sets_change(migrated: Settings) -> None:
    """Le chiffre seul se lirait comme une charge comparable d'un joueur a
    l'autre, ce qu'il n'est justement pas."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)
    corps = " ".join(build_prompt(session, settings=migrated, now=NOW).body.split())

    assert "elle ne se compare pas" in corps
    assert "passe pour un marathonien" in corps


def test_la_saison_en_cours_est_redemandee_chaque_jour(migrated: Settings) -> None:
    """Le fichier est publie une fois par semaine mais **aucun jour connu** : il
    se remplit a mesure que les tournois se terminent. Caler la relance sur notre
    propre derniere collecte manquait une publication entiere — releve le 8 aout,
    l'historique s'arretait au 3 et n'aurait ete redemande que le 13."""
    assert tennis_history.CURRENT_SEASON_TTL_HOURS == 24


def test_le_prompt_autorise_un_angle_sur_un_marche_a_relever(migrated: Settings) -> None:
    """Le premier libelle, « Non jouable », se trompait de mot et a fait le degat
    qu'il devait empecher : Betclic sert bien le handicap jeux et le total de
    jeux sur son site, c'est notre collecte qui ne les remonte pas. Une analyse
    reelle a renonce a deux angles de jeux pour se rabattre sur le vainqueur,
    alors que les paris etaient posables.

    L'affirmation est **gardee par sport** depuis qu'on sait qu'elle ne vaut pas
    partout : elle a ete verifiee au tennis, elle est fausse au football pour les
    lignes asiatiques en quart, qu'aucun book français ne propose."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)
    corps = " ".join(build_prompt(session, settings=migrated, now=NOW).body.split())

    assert "Le bookmaker propose bien ces marchés sur son site" in corps
    assert "est un marché présent" in corps
    assert "pas une raison de se rabattre sur le vainqueur" in corps
    assert "lignes en quart" not in corps, "regle de football, pas de tennis"


def test_le_prompt_reclame_un_score_exact_en_sets_au_tennis(migrated: Settings) -> None:
    """Aucune cote n'existe pour ce marche — The Odds API ne le sert pas au
    tennis, le bookmaker si. La proposition se fait donc **sans prix**, hors du
    tableau des selections et hors des combines : inventer une cote la ferait
    entrer en base et fausserait le palier comme le taux de reussite."""
    from myassistantbet.services.prompt import build_prompt

    _serie(migrated, [("6-4 6-4", True)] * 5)
    session = _lot(migrated, "tennis", "Jiri Lehecka", "Vit Kopriva", COMMENCE)
    corps = " ".join(build_prompt(session, settings=migrated, now=NOW).body.split())

    assert "Score exact en sets" in corps
    assert "Aucune cote n'existe pour ce marché" in corps
    assert "n'en fais pas une ligne de la section C" in corps


def test_le_score_exact_en_sets_ne_sort_pas_sur_un_lot_de_football(migrated: Settings) -> None:
    """Meme regle que partout : ce qui n'a pas d'objet dans le lot est omis."""
    from myassistantbet.services.prompt import build_prompt

    corps = build_prompt(_lot(migrated, "football"), settings=migrated).body

    assert "Score exact en sets" not in corps
